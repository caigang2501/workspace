import openpyxl,xlrd
import pandas as pd


def parse_xls(path):
    wb = xlrd.open_workbook(path)
    sheet_names = wb.sheet_names()
    data = []
    is_first = True
    for sheet_name in sheet_names:
        sheet = wb.sheet_by_name(sheet_name)
        for row_index in range(sheet.nrows):
            row = sheet.row_values(row_index)
            if is_first:
                is_first = False
                columns = row
            else:
                row = list(row)
                row.append(sheet_name)
                # row = [cell.value for cell in row]
                data.append(row)
        is_first = False
    columns = list(columns)
    columns.append('sheet_name')
    df = pd.DataFrame(data,columns=columns)
    return df    
import openpyxl,xlrd
import pandas as pd

def parse_xlsx(path):
    wb = openpyxl.load_workbook(path)
    sheet_names = wb.sheetnames
    data = []
    is_first = True
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):#按列遍历：sheet.iter_cols
            if is_first:
                is_first = False
                columns = row
            else:
                row = list(row)
                row.append(sheet_name)
                # row_data = [cell.value for cell in row]
                data.append(row)
        is_first = False
    wb.close()
    columns = list(columns)
    columns.append('sheet_name')
    df = pd.DataFrame(data,columns=columns)
    return df
